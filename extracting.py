from openpyxl import Workbook, load_workbook
wb=load_workbook('excel/excel_test.xlsx')
ws=wb.active
total=0
max_row=ws.max_row

for i in range(2,max_row+1):
    rate=ws['B'+str(i)].value
    hours=ws['C'+str(i)].value
    if (type(rate)!=str and type(hours)!=str):
        salary=float(rate)*float(hours)
        if salary > 3000:
            total += 1
print(total)
wb.close()
